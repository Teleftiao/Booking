import os
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# --- 1. SETTING & STYLE ---
st.set_page_config(
    page_title="Hotel Flow Grid Pro", 
    layout="wide", 
    page_icon="🏨",
    initial_sidebar_state="collapsed",  # Better for mobile
    menu_items={
        "About": "Hotel Booking Management System v1.0"
    }
)

# Mobile-friendly viewport
st.markdown("""
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    """, unsafe_allow_html=True)

st.markdown(
    """
    <style>
    .main { background-color: #0d1117; color: #c9d1d9; }
    .stMetric { background-color: #161b22; padding: 15px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.3); color: #c9d1d9; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Nama file CSV
INV_FILE = "data_inventaris.csv"
LEDGER_FILE = "data_transaksi.csv"
BOOKING_FILE = "data_booking.csv"
ROOM_FILE = "data_kamar.csv"


# --- 2. DATA ENGINE ---
def load_data():
    inv = pd.read_csv(INV_FILE) if os.path.exists(INV_FILE) else pd.DataFrame(columns=["Barang", "Stok", "Harga_Modal", "Harga_Jual"])
    leg = pd.read_csv(LEDGER_FILE) if os.path.exists(LEDGER_FILE) else pd.DataFrame(columns=["Tanggal", "Barang", "Tipe", "Jumlah", "Total_IDR"])
    book = pd.read_csv(BOOKING_FILE) if os.path.exists(BOOKING_FILE) else pd.DataFrame(columns=["Tanggal", "Customer", "No_Kamar", "Malam", "Biaya", "Status"])

    if os.path.exists(ROOM_FILE):
        rooms = pd.read_csv(ROOM_FILE)
    else:
        rooms = pd.DataFrame({"No_Kamar": [str(101 + i) for i in range(10)], "Tipe": "Standard"})

    return inv, leg, book, rooms


def save_data(inv, leg, book, rooms):
    inv.to_csv(INV_FILE, index=False)
    leg.to_csv(LEDGER_FILE, index=False)
    book.to_csv(BOOKING_FILE, index=False)
    rooms.to_csv(ROOM_FILE, index=False)


# Inisialisasi Session State
if "inventory" not in st.session_state:
    st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms = load_data()


# --- 3. LOGIK VISUAL GRID (ROOM RACK) ---
def render_hotel_grid():
    st.subheader("🗓️ Visual Room Rack (14 Hari)")

    start_date = datetime.now().date()
    date_range = [start_date + timedelta(days=x) for x in range(14)]
    date_labels = [d.strftime("%d\n%b") for d in date_range]

    room_labels = (
        st.session_state.rooms["No_Kamar"]
        .astype(str)
        .dropna()
        .drop_duplicates()
        .tolist()
    )

    grid_df = pd.DataFrame(
        "🟢 Tersedia",
        index=pd.Index(room_labels, name="Room"),
        columns=pd.Index(date_labels, name="Tanggal"),
    )

    for _, row in st.session_state.booking.iterrows():
        try:
            checkin = datetime.strptime(str(row["Tanggal"]), "%Y-%m-%d").date()
            malam = int(row["Malam"])

            for i in range(malam):
                current_date = checkin + timedelta(days=i)
                if start_date <= current_date < start_date + timedelta(days=14):
                    date_label = current_date.strftime("%d\n%b")
                    room_label = str(row["No_Kamar"])
                    if room_label in grid_df.index:
                        grid_df.at[room_label, date_label] = f"🔴 {row['Customer']}"
        except Exception:
            continue

    def style_grid(val):
        if isinstance(val, str) and "🔴" in val:
            return "background-color: #ff4b4b; color: white; font-size: 11px; font-weight: bold; text-align: center;"
        return "background-color: #2ecc71; color: white; font-size: 10px; text-align: center;"

    st.dataframe(grid_df.style.map(style_grid), use_container_width=True, height=400)


# --- 3B. MONTHLY CASHFLOW ANALYSIS ---
def generate_monthly_cashflow():
    """Generate monthly cashflow summary"""
    ledger = st.session_state.ledger.copy()
    
    if ledger.empty:
        return pd.DataFrame()
    
    # Convert Tanggal to datetime
    ledger["Tanggal"] = pd.to_datetime(ledger["Tanggal"], errors="coerce")
    ledger = ledger.dropna(subset=["Tanggal"])
    
    # Extract Year-Month
    ledger["YearMonth"] = ledger["Tanggal"].dt.to_period("M")
    
    # Separate income and expenses
    income = ledger[ledger["Tipe"].isin(["Penjualan", "Booking"])].groupby("YearMonth")["Total_IDR"].sum()
    expenses = ledger[ledger["Tipe"] == "Stok Masuk"].groupby("YearMonth")["Total_IDR"].sum()
    
    # Create monthly summary
    all_months = sorted(set(income.index) | set(expenses.index))
    monthly_data = []
    
    for month in all_months:
        monthly_data.append({
            "Bulan": str(month),
            "Pemasukan": income.get(month, 0),
            "Pengeluaran": expenses.get(month, 0),
            "Saldo": income.get(month, 0) - expenses.get(month, 0)
        })
    
    return pd.DataFrame(monthly_data)


def export_cashflow_excel():
    """Export monthly cashflow to Excel with formatting"""
    monthly_cf = generate_monthly_cashflow()
    
    if monthly_cf.empty:
        return None
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Arus Kas Bulanan"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Add title
    ws.merge_cells("A1:D1")
    ws["A1"] = "LAPORAN ARUS KAS BULANAN"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add timestamp
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Laporan hingga: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.append([])  # Empty row
    
    # Add headers
    headers = ["Bulan", "Pemasukan (Rp)", "Pengeluaran (Rp)", "Saldo (Rp)"]
    ws.append(headers)
    
    # Format header row
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    total_income = 0
    total_expense = 0
    total_balance = 0
    
    for _, row in monthly_cf.iterrows():
        ws.append([
            row["Bulan"],
            row["Pemasukan"],
            row["Pengeluaran"],
            row["Saldo"]
        ])
        total_income += row["Pemasukan"]
        total_expense += row["Pengeluaran"]
        total_balance += row["Saldo"]
        
        # Format data rows
        row_num = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
    
    # Add totals row
    ws.append([])
    total_row = ws.max_row + 1
    ws.append(["TOTAL", total_income, total_expense, total_balance])
    
    for col in range(1, 5):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    
    # Add detailed transactions sheet
    ws2 = wb.create_sheet("Transaksi Detail")
    
    # Add headers to detailed sheet
    headers_detail = ["Tanggal", "Barang", "Tipe", "Jumlah", "Total (Rp)"]
    ws2.append(headers_detail)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add all transactions
    ledger = st.session_state.ledger.copy()
    for _, row in ledger.iterrows():
        ws2.append([
            row["Tanggal"],
            row["Barang"],
            row["Tipe"],
            row["Jumlah"],
            row["Total_IDR"]
        ])
        
        row_num = ws2.max_row
        for col in range(1, 6):
            cell = ws2.cell(row=row_num, column=col)
            cell.border = border
            if col in [4, 5]:
                cell.alignment = Alignment(horizontal="right")
                if col == 5:
                    cell.number_format = '#,##0'
    
    # Adjust column widths for detailed sheet
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output



st.title("🏨 Hotel Flow Grid Pro")

cash_in = st.session_state.ledger[st.session_state.ledger["Tipe"].isin(["Penjualan", "Booking"])] ["Total_IDR"].sum()
cash_out = st.session_state.ledger[st.session_state.ledger["Tipe"] == "Stok Masuk"]["Total_IDR"].sum()
balance = cash_in - cash_out

c1, c2, c3, c4 = st.columns(4)
c1.metric("Total Pendapatan", f"Rp {cash_in:,.0f}".replace(",", "."))
c2.metric("Total Modal/Stok", f"Rp {cash_out:,.0f}".replace(",", "."))
c3.metric("Saldo Bersih", f"Rp {balance:,.0f}".replace(",", "."), delta=f"Rp {balance:,.0f}".replace(",", "."))
c4.metric("Kamar Terisi (Hari Ini)", f"{len(st.session_state.booking[st.session_state.booking['Tanggal'] == str(datetime.now().date())])} Kamar")

st.divider()

# --- 5. TAMPILAN UTAMA ---
render_hotel_grid()

st.divider()

# --- 6. MONTHLY CASHFLOW EXPORT ---
st.subheader("📊 Laporan Arus Kas Bulanan")
col_cf1, col_cf2, col_cf3 = st.columns(3)

monthly_cf = generate_monthly_cashflow()
if not monthly_cf.empty:
    col_cf1.metric("Total Pemasukan", f"Rp {monthly_cf['Pemasukan'].sum():,.0f}".replace(",", "."))
    col_cf2.metric("Total Pengeluaran", f"Rp {monthly_cf['Pengeluaran'].sum():,.0f}".replace(",", "."))
    col_cf3.metric("Total Saldo", f"Rp {monthly_cf['Saldo'].sum():,.0f}".replace(",", "."))
    
    st.dataframe(monthly_cf, use_container_width=True, hide_index=True)
    
    # Download button
    excel_file = export_cashflow_excel()
    if excel_file:
        st.download_button(
            label="📥 Download Excel - Arus Kas Bulanan",
            data=excel_file,
            file_name=f"Arus_Kas_Bulanan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("📋 Belum ada data transaksi")

st.divider()

# --- 7. INPUT & MANAJEMEN ---
tab_book, tab_inv, tab_settings = st.tabs(["📅 Reservasi Baru", "🛒 Penjualan Produk", "⚙️ Pengaturan"])

with tab_book:
    st.subheader("Input Reservasi")
    with st.form("new_booking"):
        col_b1, col_b2, col_b3 = st.columns(3)
        b_date = col_b1.date_input("Tanggal Check-in")
        b_name = col_b2.text_input("Nama Tamu")
        b_room = col_b3.selectbox("No Kamar", st.session_state.rooms["No_Kamar"].tolist())

        col_b4, col_b5, col_b6 = st.columns(3)
        b_malam = col_b4.number_input("Durasi (Malam)", min_value=1)
        b_price = col_b5.number_input("Total Bayar / DP (Rp)", min_value=0, step=50000)
        b_stat = col_b6.selectbox("Status", ["Lunas", "DP"])

        if st.form_submit_button("Simpan Reservasi"):
            new_row = pd.DataFrame(
                [{"Tanggal": str(b_date), "Customer": b_name, "No_Kamar": str(b_room), "Malam": b_malam, "Biaya": b_price, "Status": b_stat}]
            )
            st.session_state.booking = pd.concat([st.session_state.booking, new_row], ignore_index=True)

            new_cash = pd.DataFrame(
                [{"Tanggal": datetime.now().strftime("%Y-%m-%d"), "Barang": f"Kamar {b_room} ({b_name})", "Tipe": "Booking", "Jumlah": 1, "Total_IDR": b_price}]
            )
            st.session_state.ledger = pd.concat([st.session_state.ledger, new_cash], ignore_index=True)

            save_data(st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms)
            st.success("Booking Berhasil!")
            st.rerun()

with tab_inv:
    st.subheader("Input Penjualan Barang")
    with st.form("new_sale"):
        col_s1, col_s2, col_s3 = st.columns(3)
        s_type = col_s1.selectbox("Tipe", ["Penjualan", "Stok Masuk"])
        s_item = col_s2.selectbox("Produk", st.session_state.inventory["Barang"].tolist() if not st.session_state.inventory.empty else ["Kosong"])
        s_qty = col_s3.number_input("Qty", min_value=1)

        if st.form_submit_button("Proses Barang"):
            if s_item != "Kosong":
                idx = st.session_state.inventory.index[st.session_state.inventory["Barang"] == s_item][0]
                price = st.session_state.inventory.at[idx, "Harga_Jual"] if s_type == "Penjualan" else st.session_state.inventory.at[idx, "Harga_Modal"]

                if s_type == "Penjualan":
                    st.session_state.inventory.at[idx, "Stok"] -= s_qty
                else:
                    st.session_state.inventory.at[idx, "Stok"] += s_qty

                new_cash = pd.DataFrame(
                    [{"Tanggal": datetime.now().strftime("%Y-%m-%d"), "Barang": s_item, "Tipe": s_type, "Jumlah": s_qty, "Total_IDR": price * s_qty}]
                )
                st.session_state.ledger = pd.concat([st.session_state.ledger, new_cash], ignore_index=True)

                save_data(st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms)
                st.success("Stok Diperbarui!")
                st.rerun()

with tab_settings:
    st.subheader("Edit Master Data")
    edit_mode = st.radio("Pilih Data yang Ingin Diedit:", ["Daftar Kamar", "Daftar Produk/Stok", "Riwayat Booking", "Arus Kas"])

    if edit_mode == "Daftar Kamar":
        upd_rooms = st.data_editor(st.session_state.rooms, num_rows="dynamic", use_container_width=True)
        if st.button("Simpan Perubahan Kamar"):
            st.session_state.rooms = upd_rooms
            save_data(st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms)
            st.rerun()

    elif edit_mode == "Daftar Produk/Stok":
        upd_inv = st.data_editor(
            st.session_state.inventory,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Harga_Modal": st.column_config.NumberColumn(format="Rp %d"),
                "Harga_Jual": st.column_config.NumberColumn(format="Rp %d"),
            },
        )
        if st.button("Simpan Perubahan Produk"):
            st.session_state.inventory = upd_inv
            save_data(st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms)
            st.rerun()

    elif edit_mode == "Riwayat Booking":
        upd_book = st.data_editor(
            st.session_state.booking,
            num_rows="dynamic",
            use_container_width=True,
            column_config={"Biaya": st.column_config.NumberColumn(format="Rp %d")},
        )
        if st.button("Simpan Perubahan Booking"):
            st.session_state.booking = upd_book
            save_data(st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms)
            st.rerun()

    elif edit_mode == "Arus Kas":
        upd_leg = st.data_editor(
            st.session_state.ledger,
            num_rows="dynamic",
            use_container_width=True,
            column_config={"Total_IDR": st.column_config.NumberColumn(format="Rp %d")},
        )
        if st.button("Simpan Perubahan Kas"):
            st.session_state.ledger = upd_leg
            save_data(st.session_state.inventory, st.session_state.ledger, st.session_state.booking, st.session_state.rooms)
            st.rerun()
